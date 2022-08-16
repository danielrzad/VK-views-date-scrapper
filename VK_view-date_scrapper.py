from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.worksheet.cell_range import CellRange
from urllib.parse import urlparse
from requests_html import HTMLSession
from bs4 import BeautifulSoup
from pprint import pprint
from datetime import datetime
from dataclasses import make_dataclass


import yaml
import os
import time


def print_html(page_text):
    ### Just in case some debugging would be needed
    soup = BeautifulSoup(page_text, "html.parser")
    print(BeautifulSoup.prettify(soup))


def make_urls_list(data_ws, min_row, url_column):
    current_row = min_row
    urls_list = []
    while current_row <= data_ws.max_row:
        url = data_ws[f'{url_column}{current_row}'].value
        if '/wall' in url:
            data_ws.delete_rows(current_row)
            continue
        urls_list.append(url)
        current_row += 1
    return urls_list


def scrape_vk(urls_list, wait_time):
    UrlData = make_dataclass(
        'UrlData', ['views', 'date', 'add_info']
    )
    url_data_dict = {}
    for idx, url in enumerate(urls_list):
        print(idx, url)
        page_text = get_page_text(url=url, headers={'Accept-Language': 'en-US'})
        try:
            url_data_dict[url] = UrlData(
                views = None,
                date = None,
                add_info = None,
            )
            scrape_views(
                page_text=page_text, url_data=url_data_dict[url]
            )
            scrape_date(
                page_text=page_text, url_data=url_data_dict[url]
            )
        except ValueError:
            if not check_for_bad_response(
                page_text=page_text, 
                urls_list=urls_list, 
                url_data_dict=url_data_dict,
                url=url,
                url_data=url_data_dict[url],
            ):
                if 'VK | Video Ext' in page_text:
                    page_text = external_video(page_text)
                    url_data_dict[url] = UrlData(
                        views = None,
                        date = None,
                        add_info = None,
                    )
                    scrape_views(
                        page_text=page_text, url_data=url_data_dict[url]
                    )
                    scrape_date(
                        page_text=page_text, url_data=url_data_dict[url]
                    )
                else:
                    print('Not known error. Debbuging needed.')     
        time.sleep(wait_time)
    return url_data_dict


def get_page_text(url, headers):
    session = HTMLSession()
    page = session.get(url=url, headers=headers)
    page_text = page.text
    return page_text


def check_for_bad_response(page_text, urls_list, url_data_dict, url, url_data):
    ### Dictionary with possible bad responses
    bad_responses = {
        'Profile hidden': 'Bad URL, links to hidden user profile.',
        'Account deleted': 'Bad URL, links to deleted user profile.',
        '    Video deleted': 'Bad URL, video deleted.',
        'this video has been restricted by its creator': 'Bad URL, access limited by its creator.',
        '404 Not Found': 'Bad URL, page not found.',
        'Video not found.': 'Bad URL, video not found.',
    }
    for key in bad_responses:
        if key in page_text:
            print('bad_response', key)
            url_data.views = bad_responses[key]
            url_data.date = None
            url_data_dict[url] = url_data
            return True
    return False


def external_video(page_text):
    time.sleep(2)
    start = page_text.find('<div id="video_box_wrap') + len('<div id="video_box_wrap')
    end = page_text.find('" class="video_box_wrap">')
    video_id = page_text[start:end]
    url = f'https://vk.com/video{video_id}'
    headers = {'Accept-Language': 'en-US'}
    return get_page_text(url, headers)


def scrape_views(page_text, url_data):
    start = page_text.find('VideoLayerInfo__views ')
    start = start + len('VideoLayerInfo__views ') + 4
    end = page_text.find(r' views<\/div>')
    n_views = page_text[start:end]
    n_views = int(n_views.replace(',', ''))
    url_data.views = n_views


def scrape_date(page_text, url_data):
    start = page_text.find('VideoLayerInfo__date')        
    start = start + len('VideoLayerInfo__date') + 3
    date = page_text[start:start + 33]
    end = date.find(r'<\/div>')
    if end == -1:
        end = date.find('from')
        url_data.add_info = 'Video probably deleted'
    date = date[:end]
    date = datetime.strptime(date.strip(), '%d %b %Y at %I:%M %p')
    url_data.date = date


def write_urls_views_to_xlsx_file(
    data_workbook, data_ws, data_file_name, url_data_dict, 
    url_column, view_column, date_column, min_row,
    ):
    max_row = data_ws.max_row
    for row_idx in range(min_row, max_row+1):
        url = data_ws[f'{url_column}{row_idx}'].value
        data_ws[f'{view_column}{row_idx}'].value = url_data_dict[url].views
        data_ws[f'{date_column}{row_idx}'].value = url_data_dict[url].date
        if url_data_dict[url].add_info:
            color_type = PatternFill(fgColor='FFC000', fill_type = 'solid')
            data_ws[f'{url_column}{row_idx}'].fill = color_type
    data_workbook.save(f'Edited___{data_file_name}')


def main():

    ### Just a reminder in case there will be need to render the page
    # session = HTMLSession()
    # page = session.get('https://vk.com/video-211854450_456239043')
    # print(page.text)
    # page.html.render()
    # results = page.html.find('.VideoLayerInfo__views ', first=True)


    ### Backup URLs list in case more testing would be needed
    ### if this will be in use, urls_list variable with make_urls_list
    ### function call needs to be commented.
    # urls_list = [
    #     'https://vk.com/video-169661808_456246592',
    # ]


    ### Load config 
    stream = open('config.yaml', 'r')
    config = yaml.safe_load(stream)
    data_file_name = config['xlsx_file_name']
    url_column = config['URLs_column']
    view_column = config['views_column']
    date_column = config['date_column']
    min_row = config['starting_row']
    wait_time = config['wait_time']

    ### Load data file
    data_workbook = load_workbook(data_file_name)
    data_ws = data_workbook.active


    urls_list = make_urls_list(data_ws=data_ws, min_row=min_row, url_column=url_column)
    url_data_dict = scrape_vk(urls_list=urls_list, wait_time=wait_time)

    write_urls_views_to_xlsx_file(
        data_workbook=data_workbook,
        data_ws=data_ws,
        data_file_name=data_file_name, 
        url_data_dict=url_data_dict, 
        url_column=url_column,
        date_column=date_column,
        view_column=view_column,
        min_row=min_row,
    )


if __name__ == "__main__":
    main()
